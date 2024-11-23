import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# Function to get the current day's keywords from Excel
def get_keywords_from_excel(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        keywords = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value]
        wb.close()
        return keywords
    except KeyError:
        raise Exception(f"Sheet '{sheet_name}' not found in the Excel file.")
    except FileNotFoundError:
        raise Exception(f"Excel file '{file_path}' not found.")
    except Exception as e:
        raise Exception(f"An error occurred while reading keywords: {e}")

# Function to write results back to Excel
def write_results_to_excel(file_path, sheet_name, results):
    try:
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        for idx, (keyword, shortest, longest) in enumerate(results, start=2):
            sheet.cell(row=idx, column=2, value=shortest)  # Column B for shortest
            sheet.cell(row=idx, column=3, value=longest)  # Column C for longest
        wb.save(file_path)
        wb.close()
    except Exception as e:
        raise Exception(f"An error occurred while writing results to Excel: {e}")

# Function to find shortest and longest suggestions on Google
def search_google_and_find_suggestions(driver, keyword):
    try:
        driver.get("https://www.google.com")
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)

        # Wait for suggestions to load
        driver.implicitly_wait(2)
        suggestions = driver.find_elements(By.XPATH, '//ul[@role="listbox"]//span')
        suggestion_texts = [suggestion.text for suggestion in suggestions if suggestion.text]

        if suggestion_texts:
            longest = max(suggestion_texts, key=len)
            shortest = min(suggestion_texts, key=len)
        else:
            longest = shortest = "No suggestions found"

        return shortest, longest
    except Exception as e:
        print(f"An error occurred while searching for keyword '{keyword}': {e}")
        return "Error", "Error"

# Main script
def main():
    # Define Excel file and sheet
    excel_file = "keyword.xlsx"
    today = datetime.datetime.now().strftime("%A")  # Get current day name
    print(f"Processing keywords for {today}")

    try:
        # Load keywords for the current day
        keywords = get_keywords_from_excel(excel_file, today)
        if not keywords:
            print(f"No keywords found in the sheet '{today}'. Exiting.")
            return

        # Set up Selenium WebDriver
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

        results = []
        try:
            for keyword in keywords:
                shortest, longest = search_google_and_find_suggestions(driver, keyword)
                print(f"Keyword: {keyword} -> Shortest: {shortest}, Longest: {longest}")
                results.append((keyword, shortest, longest))

        finally:
            driver.quit()

        # Write results back to Excel
        write_results_to_excel(excel_file, today, results)
        print("Results written to Excel.")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
